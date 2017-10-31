import xlrd
import xlutils

file = 'output.xls'
wb = xlrd.open_workbook(file)
print(wb)
sheet = wb.sheet_by_index(0)
ncols = sheet.ncols
key = sheet.row_values(0,1,)
print(key)


# wb2 = xlwt.Workbook()
# sheet = wb2.add_sheet('sheet2')
# sheet.write(0, 1, 'hello world')
# wb.save('test2.xls')


'''
使用opnepyxl
from openpyxl import *
 # 新建文件
wb = load_workbook('test.xlsx')
ws = wb.get_sheet_by_name('Sheet2')
c=ws['A2']
print(c.value)
c.value = 'world'
print(c.value)
wb.save('test.xlsx')
'''
