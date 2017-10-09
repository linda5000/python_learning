import xlwt
wb = xlwt.Workbook()
sheet = wb.add_sheet('sheet1')
sheet.write(0, 1, 'hello world')
wb.save('2.xls')

'''
使用opnepyxl
from openpyxl import *
 # 新建文件
wb = load_workbook('test.xlsx')
ws = wb.get_sheet_by_name('Sheet2')
c=ws['A2']
print(c.value)
c.value = 'world'
print(c.value)
wb.save('test.xlsx')
'''




import cx_Oracle
tns=cx_Oracle.makedsn('10.10.100.234',1521,'wlzy')
db=cx_Oracle.connect('sm','sm',tns)
cr=db.cursor()

pr = {'zc_no':'000013008605'}
sql='select * from zc_sm where ASSETSCARDCODE =:zc_no'
cr.execute(sql,pr)
rs=cr.fetchall()
print(rs)
for i in rs:
    print(i)
cr.close()
db.close()
